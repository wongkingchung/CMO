# -*- coding: utf-8 -*-
import codecs
import pyodbc
import openpyxl
import ConfigParser

import os
import glob
import shutil
import datetime
from sendemail import sendmail

config = ConfigParser.ConfigParser()
config.read('configloc.ini')

configpath = config.get('CONFIG','path')

config.read(configpath+"config.ini")

server = config.get('DB','server')
database = config.get('DB','database')
username = config.get('DB','username')
password = config.get('DB','password')
schema = config.get('DB','schema')

promofilepath = config.get('PROMOTION','file')
logpath = config.get('PROMOTION','logpath')

os.chdir(promofilepath)
promofiles = glob.glob("[a-z]*.xlsx")
try:
    promofile =  promofilepath+promofiles[0]
except IndexError:
    os._exit(0)
else:
    print "continue", promofile



driver= '{ODBC Driver 13 for SQL Server}'
cnxn = pyodbc.connect('DRIVER='+driver+';PORT=1433;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+';PWD='+ password+';Charset=UTF8')
cursor = cnxn.cursor()



wb = openpyxl.load_workbook(promofile)
ws = wb['Client Promotion']
ended = False
r=1
c=0

SQLCommand = ("DELETE "+schema+".imp_promotion;")
cursor.execute(SQLCommand)

try:
    while not ended:
        r=r+1
        planid = ws.cell(row=r, column=3).value
        promotion_text = ws.cell(row=r, column=5).value
        promotion_link = ""
        if ws.cell(row=r, column=5).hyperlink != None:
            promotion_link = ws.cell(row=r, column=5).hyperlink.target
        active = ws.cell(row=r, column=6).value
        language = ws.cell(row=r, column=7).value
        startdate = ws.cell(row=r, column=8).value
        enddate = ws.cell(row=r, column=9).value

        if enddate == None:
            enddate = '2099-12-31'

        if active == None:
            active = 1

        if planid == None and promotion_text == None and  startdate == None:
            ended = True
        else:
            if planid == None or promotion_text == None or startdate == None:
                continue
        
            if promotion_link != "":
                promotion_text = promotion_text + '<a href="'+promotion_link+'"> (click here)</a>' 

            if language == None:
                SQLCommand = u"INSERT INTO "+schema+".imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','en','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
                cursor.execute(SQLCommand)
                SQLCommand = u"INSERT INTO "+schema+".imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','zh','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
                cursor.execute(SQLCommand)
            else:
                SQLCommand = u"INSERT INTO "+schema+".imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','"+language+"','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
                cursor.execute(SQLCommand)

    #        print SQLCommand
except Exception, error:
    sendmail('There are errors in importing sheet Client Promotion. Please contact IT. ' + str(error) )
    os._exit(0)
    

ws = wb['BV Promotion']
ended = False
r=1
c=0

try:
    while not ended:
        r=r+1
        planid = ws.cell(row=r, column=3).value
        promotion_text = ws.cell(row=r, column=4).value
        promotion_link = ""
        if ws.cell(row=r, column=4).hyperlink != None:
            promotion_link = ws.cell(row=r, column=4).hyperlink.target

    #    active = ws.cell(row=r, column=5).value
    #    startdate = ws.cell(row=r, column=6).value
    #    enddate = ws.cell(row=r, column=7).value
        active = 1
        startdate = ws.cell(row=r, column=5).value
        enddate = ws.cell(row=r, column=6).value

        if enddate == None:
            enddate = '2099-12-31'

        if active == None:
            active = 1

        if planid == None and promotion_text == None and  startdate == None:
            ended = True
        else:
            if planid == None or promotion_text == None or startdate == None:
                continue

            if promotion_link != "":
                promotion_text = promotion_text + '<a href="'+promotion_link+'"> (click here)</a>' 

            SQLCommand = u"INSERT INTO "+schema+".imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','en','bv',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
            print SQLCommand
            cursor.execute(SQLCommand)
            SQLCommand = u"INSERT INTO "+schema+".imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','zh','bv',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
            print SQLCommand
            cursor.execute(SQLCommand)

except Exception, error:
    sendmail('There are errors in importing sheet BV Promotion. Please contact IT. ' + str(error) )
    os._exit(0)



## insert new promotions
#SQLCommand = u"insert into icomparev2.convoy_feature  select plan_id, case when type='cp' then 'promotion' else 'bv_promotion' end as feature_name, stuff((select '<br><br>' +  (promotion) from icomparev2.imp_promotion where plan_id = a.plan_id and language = a.language and active = a.active order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, 'D' as display_type, 'text' as data_type, active, getdate(), language, min(startdate) as startdate , max(enddate) as enddate from icomparev2.imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 and not exists (select 1 from icomparev2.convoy_feature f where f.plan_id = a.plan_id and f.language = a.language ) group by plan_id, language,active, type"
SQLCommand = u"insert into "+schema+".convoy_feature  select plan_id, case when type='cp' then 'promotion' else 'bv_promotion' end as feature_name, stuff((select '<br><br>' +  (promotion) from "+schema+".imp_promotion where plan_id = a.plan_id and language = a.language and type = a.type and active = a.active order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, 'D' as display_type, 'text' as data_type, active, getdate(), language, min(startdate) as startdate , max(enddate) as enddate from "+schema+".imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 and not exists (select 1 from "+schema+".convoy_feature f where f.plan_id = a.plan_id and f.language = a.language and f.feature_name = case when a.type ='cp' then 'promotion' else 'bv_promotion' end ) group by plan_id, language,active, type"
#print SQLCommand
cursor.execute(SQLCommand)

## inactivate promotions
SQLCommand = u"update "+schema+".convoy_feature set is_active = 0 , display_start_date = ip.startdate, display_end_date = ip.enddate from ( select * from "+schema+".imp_promotion where active = 0 or not (getdate() >= startdate and getdate() <= enddate)) ip where "+schema+".convoy_feature.plan_id = ip.plan_id  and "+schema+".convoy_feature.feature_name = case when ip.type = 'cp' then 'promotion' else 'bv_promotion' end"
##print SQLCommand
cursor.execute(SQLCommand)

## update promotions
SQLCommand = u"update "+schema+".convoy_feature set feature_value = ip.promotion, display_start_date = ip.startdate, display_end_date = ip.enddate, is_active=1 from (select plan_id, type, stuff((select '<br><br>' +  (promotion) from "+schema+".imp_promotion   where plan_id = a.plan_id   and language = a.language   and active = a.active and type = a.type and getdate()>=startdate and getdate() <=enddate and active = 1 order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, language,  min(startdate) as startdate , max(enddate) as enddate from "+schema+".imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 group by plan_id, language,active, type ) ip where "+schema+".convoy_feature.plan_id = ip.plan_id and "+schema+".convoy_feature.language = ip.language and "+schema+".convoy_feature.feature_name = case when ip.type = 'cp' then 'promotion' else 'bv_promotion' end"
##print SQLCommand
cursor.execute(SQLCommand)

cnxn.commit()
cursor.close()


if logpath[-1] != '\\':
    logpath = logpath + '\\'

logfile = logpath + 'log_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
#shutil.copy(promofile, logfile)
shutil.move(promofile, logfile)

sendmail('The promotion file '+promofile+ ' has been uploaded.')
