# -*- coding: utf-8 -*-
import codecs
import pyodbc
import openpyxl
import ConfigParser

import os
import shutil
import datetime


config = ConfigParser.ConfigParser()
config.read('config.ini')

server = config.get('DB','server')
database = config.get('DB','database')
username = config.get('DB','username')
password = config.get('DB','password')

promofile = config.get('PROMOTION','file')
logpath = config.get('PROMOTION','logpath')


driver= '{ODBC Driver 13 for SQL Server}'
cnxn = pyodbc.connect('DRIVER='+driver+';PORT=1433;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+';PWD='+ password+';Charset=UTF8')
cursor = cnxn.cursor()



wb = openpyxl.load_workbook(promofile)
ws = wb['Client Promotion']
ended = False
r=1
c=0

SQLCommand = ("DELETE icompare_uat.imp_promotion;")
cursor.execute(SQLCommand)


while not ended:
    r=r+1
    planid = ws.cell(row=r, column=3).value
    promotion_text = ws.cell(row=r, column=5).value
    promotion_link = ""
    if ws.cell(row=r, column=5).hyperlink != None:
        promotion_link = ws.cell(row=r, column=5).hyperlink.target
    active = ws.cell(row=r, column=6).value
    language = ws.cell(row=r, column=7).value
    startdate = ws.cell(row=r, column=8).value
    enddate = ws.cell(row=r, column=9).value

    if enddate == None:
        enddate = '2099-12-31'

    if active == None:
        active = 1

    if planid == None and promotion_text == None and  startdate == None:
        ended = True
    else:
        if planid == None or promotion_text == None or startdate == None:
            continue
    
        if promotion_link != "":
            promotion_text = promotion_text + '<a href="'+promotion_link+'"> (click here)</a>' 

        if language == None:
            SQLCommand = u"INSERT INTO icompare_uat.imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','en','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
            cursor.execute(SQLCommand)
            SQLCommand = u"INSERT INTO icompare_uat.imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','zh','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
            cursor.execute(SQLCommand)
        else:
            SQLCommand = u"INSERT INTO icompare_uat.imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','"+language+"','cp',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
            cursor.execute(SQLCommand)

        print SQLCommand


ws = wb['BV Promotion']
ended = False
r=1
c=0

while not ended:
    r=r+1
    planid = ws.cell(row=r, column=3).value
    promotion_text = ws.cell(row=r, column=4).value
    promotion_link = ""
    if ws.cell(row=r, column=4).hyperlink != None:
        promotion_link = ws.cell(row=r, column=4).hyperlink.target

    active = ws.cell(row=r, column=5).value
    startdate = ws.cell(row=r, column=6).value
    enddate = ws.cell(row=r, column=7).value

    if enddate == None:
        enddate = '2099-12-31'

    if active == None:
        active = 1

    if planid == None and promotion_text == None and  startdate == None:
        ended = True
    else:
        if planid == None or promotion_text == None or startdate == None:
            continue

        if promotion_link != "":
            promotion_text = promotion_text + '<a href="'+promotion_link+'"> (click here)</a>' 

        SQLCommand = u"INSERT INTO icompare_uat.imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','en','bv',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
        print SQLCommand
        cursor.execute(SQLCommand)
        SQLCommand = u"INSERT INTO icompare_uat.imp_promotion (plan_id, promotion, language, type, active, startdate, enddate) VALUES (" +str(planid)+",N'" + promotion_text+"','zh','bv',"+str(active)+",'"+str(startdate)+"','"+str(enddate)+"');"
        print SQLCommand
        cursor.execute(SQLCommand)

## insert new promotions
SQLCommand = u"insert into icompare_uat.convoy_feature  select plan_id, 'promotion' as feature_name, stuff((select '<br><br>' +  (promotion) from icompare_uat.imp_promotion where plan_id = a.plan_id and language = a.language and active = a.active order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, 'D' as display_type, 'text' as data_type, active, getdate(), language, min(startdate) as startdate , max(enddate) as enddate from icompare_uat.imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 and not exists (select 1 from icompare_uat.convoy_feature f where f.plan_id = a.plan_id and f.language = a.language ) group by plan_id, language,active"
print SQLCommand
cursor.execute(SQLCommand)

## inactivate promotions
SQLCommand = u"update icompare_uat.convoy_feature set is_active = 0 from ( select * from icompare_uat.imp_promotion where active = 0 or not (getdate() >= startdate and getdate() <= enddate)) ip where icompare_uat.convoy_feature.plan_id = ip.plan_id  and icompare_uat.convoy_feature.feature_name ='promotion'"
print SQLCommand
cursor.execute(SQLCommand)

## update promotions
SQLCommand = u"update icompare_uat.convoy_feature set feature_value = ip.promotion, display_start_date = ip.startdate, display_end_date = ip.enddate, is_active=1 from (select plan_id, stuff((select '<br><br>' +  (promotion) from icompare_uat.imp_promotion   where plan_id = a.plan_id   and language = a.language   and active = a.active  and getdate()>=startdate and getdate() <=enddate and active = 1 order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, language,  min(startdate) as startdate , max(enddate) as enddate from icompare_uat.imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 group by plan_id, language,active ) ip where icompare_uat.convoy_feature.plan_id = ip.plan_id and icompare_uat.convoy_feature.language = ip.language and icompare_uat.convoy_feature.feature_name = 'promotion'"
print SQLCommand
cursor.execute(SQLCommand)

cnxn.commit()

if logpath[-1] != '\\':
    logpath = logpath + '\\'

logfile = logpath + 'log_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
shutil.copy(promofile, logfile)
