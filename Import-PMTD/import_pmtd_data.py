# -*- coding: utf-8 -*-
import codecs
import pyodbc
import openpyxl
import ConfigParser
import sys

import os
import glob
import shutil
import datetime
from sendemail import sendmail

config = ConfigParser.ConfigParser()
config.read('configloc.ini')

configpath = config.get('CONFIG','path')

config.read(configpath+"config.ini")


server = config.get('DB','server')
database = config.get('DB','database')
username = config.get('DB','username')
password = config.get('DB','password')
schema = config.get('DB','schema')

pmtdfilepath = config.get('PMTD','file')
logpath = config.get('PMTD','logpath')


sheet = '2. PMTD Data'
sheet2 = '3. Rider Data'

os.chdir(pmtdfilepath)
pmtdfiles = glob.glob("[a-z]*.xlsx")
try:
    i = 0
    chknextfile = True
    while chknextfile:
        pmtdfile =  pmtdfilepath+pmtdfiles[i]
        wb = openpyxl.load_workbook(pmtdfile, data_only=True)
        if sheet in wb.sheetnames:
            if sheet2 in wb.sheetnames:
                chknextfile = False

        if chknextfile:
            i = i + 1
            wb.close()
            
except IndexError:
    os._exit(0)
else:
    print "continue", pmtdfile
    


driver= '{ODBC Driver 13 for SQL Server}'
cnxn = pyodbc.connect('DRIVER='+driver+';PORT=1433;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+';PWD='+ password)
cursor = cnxn.cursor()

SQLCommand = ("DELETE "+schema+".import_pmtd_data;")
cursor.execute(SQLCommand)


def readExcelSheet(sheet, hr ):
    ws = wb[sheet]
    col_ended = False
    row_ended = False
    r = hr

    SQLCommand = "INSERT INTO "+schema+".import_pmtd_data (row, fieldname, fieldvalue, sheet) VALUES (?,?,?,?)"

    try:
        while not row_ended:
            r = r + 1
            c=1
            fieldvalue = ws.cell(row=r, column=c).value
            if fieldvalue == None:
                row_ended = True
            else:
                col_ended = False
                
            while not col_ended:
                row = r - hr
                fieldname = ws.cell(row=hr, column=c).value
                fieldvalue = ws.cell(row=r, column=c).value 
                c = c + 1

                if fieldname == None:
                    col_ended = True
                else:
                    print row, fieldname#, fieldvalue, sheet
                    cursor.execute(SQLCommand, row, fieldname, fieldvalue, sheet)
            
    except Exception, error:
        sendmail('There are errors in importing sheet ' + sheet + '. Please contact IT. ' + str(error) + ';row='  + str(row) + '; fieldname='+ fieldname )
        print 'There is an error', str(error) , 'row',str(row), 'fieldname',fieldname, 'fieldvalue',fieldvalue
        os._exit(0)
        
#    SQLCommand = "exec " + schema +".sp_update_pmtd_data"
#    SQLCommand = "exec " + schema + sp
#    cursor.execute(SQLCommand)

readExcelSheet(sheet,2,)
readExcelSheet(sheet2,1)

SQLCommand = "exec " + schema +".sp_update_pmtd_data"
cursor.execute(SQLCommand)
               

cnxn.commit()
cursor.close()

wb.close()

if logpath[-1] != '\\':
    logpath = logpath + '\\'

logfile = logpath + 'log_pmtd_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
shutil.move(pmtdfile, logfile)

sendmail('The PMTD file '+pmtdfile+ ' has been uploaded.')


    

