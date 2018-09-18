# -*- coding: utf-8 -*-
import codecs
import pyodbc
import openpyxl
import ConfigParser

import os
import shutil
import datetime


config = ConfigParser.ConfigParser()
config.read('config.ini')

server = config.get('DB','server')
database = config.get('DB','database')
username = config.get('DB','username')
password = config.get('DB','password')

promofile = config.get('PROMOTION','file')
logpath = config.get('PROMOTION','logpath')


driver= '{ODBC Driver 13 for SQL Server}'
cnxn = pyodbc.connect('DRIVER='+driver+';PORT=1433;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+';PWD='+ password+';Charset=UTF8')
cursor = cnxn.cursor()



wb = openpyxl.load_workbook(promofile)
ws = wb['Sheet1']
ended = False
r=1
c=0

SQLCommand = ("DELETE icompare.import_vsmart_bv;")
cursor.execute(SQLCommand)


while not ended:
    r=r+1
    icompare_id = ws.cell(row=r, column=1).value
    vsmart_id = ws.cell(row=r, column=2).value

    if vsmart_id == None:
        ended = True
        continue
    
    term_type = ws.cell(row=r, column=5).value
    if term_type == "Payment":
        term_value = ws.cell(row=r, column=6).value
    else:
        term_value = ws.cell(row=r, column=8).value

    if term_type == None:
        term_type = ""

    if term_value == None:
        term_value = 0
        
    tier = ws.cell(row=r, column=9).value
    if tier == None:
        tier = ""
        
    installment = ws.cell(row=r, column=10).value
    if installment == None:
        installment = ""
        
    bv_factor = ws.cell(row=r, column=16).value
    if bv_factor == None:
        bv_factor = 0
        
    effective_from = ws.cell(row=r, column=20).value
    if effective_from == None:
        effective_from = "1900-01-01"
        
    effective_to = ws.cell(row=r, column=21).value
    if effective_to == None:
        effective_to = "2999-01-01" 

    SQLCommand = u"INSERT INTO icompare.import_vsmart_bv (icompare_id, vsmart_id, term_type, term_value, tier, installment, bv_factor, effective_from, effective_to) VALUES ("
    SQLCommand = SQLCommand  + str(icompare_id)+",'" + vsmart_id +"','" + term_type + "'," + str(term_value) + ",'" + str(tier) + "','" + installment +"'," + str(bv_factor) + ",'"  +str(effective_from)+"','"+str(effective_to)+"');"
    print SQLCommand
    cursor.execute(SQLCommand)

 


## insert new promotions
##SQLCommand = u"insert into icompare_uat.convoy_feature  select plan_id, 'promotion' as feature_name, stuff((select '<br><br>' +  (promotion) from icompare_uat.imp_promotion where plan_id = a.plan_id and language = a.language and active = a.active order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, 'D' as display_type, 'text' as data_type, active, getdate(), language, min(startdate) as startdate , max(enddate) as enddate from icompare_uat.imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 and not exists (select 1 from icompare_uat.convoy_feature f where f.plan_id = a.plan_id and f.language = a.language ) group by plan_id, language,active"
##print SQLCommand
##cursor.execute(SQLCommand)

## inactivate promotions
##SQLCommand = u"update icompare_uat.convoy_feature set is_active = 0 from ( select * from icompare_uat.imp_promotion where active = 0 or not (getdate() >= startdate and getdate() <= enddate)) ip where icompare_uat.convoy_feature.plan_id = ip.plan_id  and icompare_uat.convoy_feature.feature_name ='promotion'"
##print SQLCommand
##cursor.execute(SQLCommand)

## update promotions
##SQLCommand = u"update icompare_uat.convoy_feature set feature_value = ip.promotion, display_start_date = ip.startdate, display_end_date = ip.enddate, is_active=1 from (select plan_id, stuff((select '<br><br>' +  (promotion) from icompare_uat.imp_promotion   where plan_id = a.plan_id   and language = a.language   and active = a.active  and getdate()>=startdate and getdate() <=enddate and active = 1 order by type desc FOR XML PATH(''),type).value('(./text())[1]','nvarchar(max)'),1,8,'') as promotion, language,  min(startdate) as startdate , max(enddate) as enddate from icompare_uat.imp_promotion a where getdate()>=startdate and getdate() <=enddate and active = 1 group by plan_id, language,active ) ip where icompare_uat.convoy_feature.plan_id = ip.plan_id and icompare_uat.convoy_feature.language = ip.language and icompare_uat.convoy_feature.feature_name = 'promotion'"
##print SQLCommand
##cursor.execute(SQLCommand)

cnxn.commit()

