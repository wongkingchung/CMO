import codecs
import openpyxl
import ConfigParser
import smtplib



config = ConfigParser.ConfigParser()
config.read('configloc.ini')

configpath = config.get('CONFIG','path')

config.read(configpath+"config.ini")



smtpserver = config.get('SMTP','server')
smtpport = config.get('SMTP','port')
smtplogin = config.get('SMTP','login')
smtppassword = config.get('SMTP','password')
receiver = config.get('SMTP','receiver')
            
print smtpserver, smtpport

#server = smtplib.SMTP_SSL(smtpserver, smtpport)
server = smtplib.SMTP(smtpserver, smtpport)
server.ehlo()
#Next, log in to the server
server.login(smtplogin, smtppassword)

#ClientSheet = config.get('PROMOTION','ClientSheet')
#ClientSheetFormat = config.get('PROMOTION','ClientSheetFormat')
#promofile = config.get('PROMOTION','file')


from email.mime.text import MIMEText
 
##//A subclass of MIMEBase, this is an intermediate base class for MIME messages that are multipart
from email.mime.multipart import MIMEMultipart

def sendmail(body):

    email_sender = 'noreply@convoy.com.hk'
    email_receiver = receiver ##'kc.wong@convoy.com.hk'
     
    subject = 'python!'
     
    msg = MIMEMultipart() ##//used for define multipart message
    msg['From'] = email_sender
    msg['To'] = email_receiver
    msg['Subject']= 'PMTD Import'
     
##    body = 'hi everyone ! this email is from python'
    msg.attach(MIMEText(body, 'plain'))  ##// attach body to the message, here email is plain so email type plain is used
    text = msg.as_string()  ##// used for converting object into plain text string
     
    server.sendmail(email_sender, email_receiver, text)

sendmail('this is a test3')

##
##def checkvalidity(Sheet):
##    ClientSheet = config.get('PROMOTION',Sheet)
##    ClientSheetFormat = config.get('PROMOTION',Sheet+'Format')
##    promofile = config.get('PROMOTION','file')
##
##    wb = openpyxl.load_workbook(promofile)
##    ws = wb[ClientSheet]
##
##    col=0
##    ClientSheetFormatFields = ClientSheetFormat.split(',')
##    for f in ClientSheetFormatFields:
##        col = col + 1
###        f1 = f.decode('utf-8').strip()
##        f1 = f.strip()
##        print f1
##        if type(ws.cell(row=1, column = col).value) == type(None):
##            f2 = ""
##        else:
###            f2 = ws.cell(row=1, column = col).value.decode('utf-8').strip()
##            f2 = ws.cell(row=1, column = col).value.strip()
##        print Sheet, f1, f2
##        if f1 != f2:
##            print ('wrong format', f1, f2)
##            msg = """From: noreply@convoy.com.hk
##To : kc.wong@convoy.com.hk\n
##Subject : PMTD Import\n
##wrong format
##"""
##            server.sendmail("kc.wong@convoy.com.hk", "kc.wong@convoy.com.hk", msg)
##            break
##
##    if col == len(ClientSheetFormatFields):            
##        print 'right format'
##        msg = """From: noreply@convoy.com.hk
##To : kc.wong@convoy.com.hk\n
##Subject : PMTD Import\n
##right format
##"""
##        server.sendmail("kc.wong@convoy.com.hk", "kc.wong@convoy.com.hk", 'right format')
##
###checkvalidity('ClientSheet')
##checkvalidity('BVSheet')
##
####Send the mail
###msg = "Hello!" # The /n separates the message from the headers
###server.sendmail("wongkingchung@gmail.com", "wongkingchung@gmail.com", msg)
